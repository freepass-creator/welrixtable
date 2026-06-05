# v5.5 엑셀 차량DB → vehicles.json 자동 동기화
# 두 파일의 trim 매칭 후 가격/잔가율 차이 출력 + 새 vehicles.v55.json 생성
# 사용법:
#   python scripts/sync-vehicles-from-excel.py
#   (검증 모드 — diff만 출력)
#   python scripts/sync-vehicles-from-excel.py --apply
#   (vehicles.json 덮어쓰기)
import sys, io, json, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

EXCEL = '_source/견적기/웰릭스모빌리티 신차 견적(20260604)_v5.5_배포용.xlsx'
JSON_PATH = 'public/data/vehicles.json'
APPLY = '--apply' in sys.argv

wb = openpyxl.load_workbook(EXCEL, data_only=True)
ws = wb['차량DB']

# 컬럼 매핑 (헤더 row 2)
COLS = {'brand':1,'model':2,'name':3,'trim':4,'price':5,'disp':6,'fuel':7,
        'tax_exempt':8,'group':9,'multi_seat':10,'r24':11,'r36':12,'r48':13,
        'r60':14,'strategic':15,'buyback_apply':16}

excel_trims = []
for r in range(3, ws.max_row + 1):
    row = {k: ws.cell(row=r, column=c).value for k, c in COLS.items()}
    if not row['brand'] or not row['trim']:
        continue
    if row['price'] is None or row['price'] == 0:
        continue
    # 가격은 만원 단위로 들어옴 — 원 단위로 (필요 시)
    p = row['price']
    if p < 1000000:  # 만원 단위
        p = p * 10000
    row['price'] = int(p)
    # 잔가율 정규화
    for k in ['r24','r36','r48','r60']:
        if row[k] is not None:
            row[k] = round(row[k], 4)
    excel_trims.append(row)

# 현재 vehicles.json 로드
with open(JSON_PATH, 'r', encoding='utf-8') as f:
    json_trims = json.load(f)

# trim string 정규화 — 매칭용
def norm(s):
    if s is None: return ''
    return str(s).replace(' ', '').replace('　', '')

def make_key_json(t):
    # 엑셀 D 컬럼 형식 = name + ' ' + 옵션상세 — vehicles.json name 사용
    trim = (t.get('trim') or '').strip()
    name = (t.get('name') or '').strip()
    # trim이 이미 name으로 시작하면 그대로, 아니면 name 추가
    if name and trim.startswith(name):
        composite = trim
    else:
        composite = f'{name} {trim}' if name else trim
    return (norm(t.get('brand') or ''), norm(composite))

def make_key_excel(t):
    # 엑셀: brand + D (세부모델 = name + trim 형식)
    return (norm(t.get('brand') or ''), norm(t.get('trim') or ''))

# 매칭
excel_by_key = {make_key_excel(t): t for t in excel_trims}
json_by_key = {make_key_json(t): t for t in json_trims}

only_excel = [k for k in excel_by_key if k not in json_by_key]
only_json = [k for k in json_by_key if k not in excel_by_key]
common = [k for k in excel_by_key if k in json_by_key]

# 공통 trim에서 가격/잔가율 차이 추출
diffs = []
COMPARE_FIELDS = ['price', 'r24', 'r36', 'r48', 'r60']
for key in common:
    e = excel_by_key[key]
    j = json_by_key[key]
    field_diffs = {}
    for f in COMPARE_FIELDS:
        ev = e.get(f)
        jv = j.get(f)
        if ev is None or jv is None:
            continue
        if isinstance(ev, float) and isinstance(jv, float):
            if abs(ev - jv) > 0.0001:
                field_diffs[f] = (jv, ev)
        elif ev != jv:
            field_diffs[f] = (jv, ev)
    if field_diffs:
        diffs.append((key, j['trim'], field_diffs))

print(f'=== vehicles.json ↔ v5.5 차량DB 비교 ===')
print(f'엑셀: {len(excel_trims)} trims / json: {len(json_trims)} trims')
print(f'매칭: {len(common)} / json에만: {len(only_json)} / 엑셀에만: {len(only_excel)}')
print()
print(f'=== 가격/잔가율 변경 ({len(diffs)} trims) ===')
for key, trim, fd in diffs[:30]:
    parts = []
    for f, (old, new) in fd.items():
        if f == 'price':
            parts.append(f'price {int(old):,} → {int(new):,}')
        else:
            parts.append(f'{f} {old} → {new}')
    print(f'  [{trim[:50]}] {" / ".join(parts)}')
if len(diffs) > 30:
    print(f'  ... 외 {len(diffs) - 30}건')

if only_excel:
    print(f'\n=== 엑셀에만 있는 trim ({len(only_excel)}건) — vehicles.json 추가 필요 ===')
    for k in only_excel[:20]:
        e = excel_by_key[k]
        print(f'  [{e["brand"]}] {e["trim"]} | 가격 {e.get("price"):,}')

if only_json:
    print(f'\n=== json에만 있는 trim ({len(only_json)}건) — 단종? ===')
    for k in only_json[:20]:
        j = json_by_key[k]
        print(f'  [{j["brand"]}] {j["trim"]}')

# --apply: vehicles.json 덮어쓰기
if APPLY and diffs:
    for key, _, fd in diffs:
        j = json_by_key[key]
        for f, (_, new_val) in fd.items():
            j[f] = new_val
    # 백업
    backup = JSON_PATH + '.v45.bak'
    if not os.path.exists(backup):
        import shutil
        shutil.copy(JSON_PATH, backup)
        print(f'\n→ 백업: {backup}')
    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(json_trims, f, ensure_ascii=False, indent=2)
    print(f'→ vehicles.json {len(diffs)}건 업데이트 완료')
else:
    print(f'\n(--apply 옵션으로 vehicles.json 자동 업데이트 가능)')
