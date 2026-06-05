# 엑셀 견적기 두 버전 직접 비교 — 변경된 셀만 자동 추출
# 사용법:
#   python scripts/compare-excel-versions.py [old.xlsx] [new.xlsx]
# 기본: v4.5 vs v5.5
#
# 출력: 시트별로 (셀좌표, 이전값 → 새값) 차이만
# 동일 셀에서 값이 다른 것만 표시. 수식이 아닌 단순 상수 차이를 우선.

import sys, io, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

old_path = sys.argv[1] if len(sys.argv) > 1 else '_source/견적기/웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx'
new_path = sys.argv[2] if len(sys.argv) > 2 else '_source/견적기/웰릭스모빌리티 신차 견적(20260604)_v5.5_배포용.xlsx'

print(f'== OLD: {os.path.basename(old_path)} ==')
print(f'== NEW: {os.path.basename(new_path)} ==\n')

# data_only=True → 계산된 결과값으로 비교
wb_old = openpyxl.load_workbook(old_path, data_only=True)
wb_new = openpyxl.load_workbook(new_path, data_only=True)

# 입력값 따라 다 변하는 시트는 노이즈 → 핵심 상수 시트만 비교
INPUT_NOISY_SHEETS = {'견적서_고객', '견적1', '견적2', '견적3', '고객카드'}
ARG_ALL = '--all' in sys.argv
common_sheets = [s for s in wb_old.sheetnames if s in wb_new.sheetnames
                 and (ARG_ALL or s not in INPUT_NOISY_SHEETS)]
only_old = [s for s in wb_old.sheetnames if s not in wb_new.sheetnames]
only_new = [s for s in wb_new.sheetnames if s not in wb_old.sheetnames]

if only_old: print(f'! OLD 에만 있는 시트: {only_old}')
if only_new: print(f'! NEW 에만 있는 시트: {only_new}')

total_diffs = 0
sheet_diffs = {}

def is_numeric(v):
    if v is None: return False
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False

def values_equal(a, b):
    if a is None and b is None: return True
    if a is None or b is None: return False
    if is_numeric(a) and is_numeric(b):
        return abs(float(a) - float(b)) < 0.0001
    return str(a).strip() == str(b).strip()

for sheet in common_sheets:
    ws_old = wb_old[sheet]
    ws_new = wb_new[sheet]
    max_r = max(ws_old.max_row, ws_new.max_row)
    max_c = max(ws_old.max_column, ws_new.max_column)
    diffs = []
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v_old = ws_old.cell(row=r, column=c).value
            v_new = ws_new.cell(row=r, column=c).value
            if values_equal(v_old, v_new):
                continue
            coord = ws_old.cell(row=r, column=c).coordinate
            # 옆에 라벨 셀 같이 (B 컬럼 또는 같은 행의 첫 텍스트 셀)
            label = ''
            for lc in range(max(1, c - 2), c):
                lv = ws_new.cell(row=r, column=lc).value
                if isinstance(lv, str) and lv.strip():
                    label = lv.strip()
                    break
            diffs.append((coord, label, v_old, v_new))
    if diffs:
        sheet_diffs[sheet] = diffs
        total_diffs += len(diffs)

# 출력 — 시트별
for sheet, diffs in sheet_diffs.items():
    print(f'\n===== [{sheet}] {len(diffs)}건 =====')
    # 숫자 차이 우선
    numeric_first = sorted(diffs, key=lambda d: 0 if (is_numeric(d[2]) and is_numeric(d[3])) else 1)
    for coord, label, v_old, v_new in numeric_first[:30]:
        old_s = str(v_old)[:30] if v_old is not None else '(빈)'
        new_s = str(v_new)[:30] if v_new is not None else '(빈)'
        label_s = f' [{label[:25]}]' if label else ''
        print(f'  {coord}{label_s}: {old_s} → {new_s}')
    if len(diffs) > 30:
        print(f'  ... 외 {len(diffs) - 30}건')

print(f'\n=== 총 차이 {total_diffs}건 ===')
print('숫자 변경 셀이 calc.js DEFAULT_CFG 변경 후보')
