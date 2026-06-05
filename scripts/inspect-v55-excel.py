# v5.5 엑셀 모든 시트 / 셀 값 dump → /tmp/v55-dump.txt
import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
import openpyxl

PATH = '_source/견적기/웰릭스모빌리티 신차 견적(20260604)_v5.5_배포용.xlsx'

# data_only=True 로 계산된 값
wb = openpyxl.load_workbook(PATH, data_only=True)
print('=== Sheet names ===')
for s in wb.sheetnames:
    print(f'  {s}')

# 각 시트 dump (값 있는 셀만)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=False))
    has_content = False
    print(f'\n===== [{sheet}] {ws.max_row}x{ws.max_column} =====')
    for row in rows[:40]:
        line = []
        for cell in row[:15]:
            v = cell.value
            if v is None:
                line.append('')
            else:
                s = str(v).replace('\n', ' ')[:50]
                line.append(s)
        if any(line):
            print('  ' + ' | '.join(line))
            has_content = True
    if not has_content:
        print('  (empty)')

# 핵심 계산 시트 — 회사/단가, 차종DB, 견적_표지 가 있을 듯
# 잔존가/수수료/이자율 같은 핵심 상수 찾기
print('\n\n=== 핵심 키워드 매칭 ===')
KEYWORDS = ['잔존', '수수료', '이자', '마진', '회수', '보험', '정비', '5.5', '5.0', 'V5.5', '버전', 'version', '대여료', '월 납입', '월 부담', '서비스 단가']
for sheet in wb.sheetnames:
    ws = wb[sheet]
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            sv = str(v)
            if any(kw in sv for kw in KEYWORDS):
                # 옆 셀(다음 컬럼)도 같이 출력 — 값 매핑
                next_cell = ws.cell(row=cell.row, column=cell.column + 1)
                print(f'  [{sheet} {cell.coordinate}] {sv[:70]}  →  {str(next_cell.value)[:30] if next_cell.value is not None else ""}')
