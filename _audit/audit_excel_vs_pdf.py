"""
Excel (차량DB) vs PDF (정상가 5%) — 완전 일치 감사.

규칙: Excel의 모든 가격이 PDF text 에 정확히 등장해야 함 (정상가 5%).
출처: 엑셀 시트는 정상가 기준 (개소세 5%), PDF의 첫 번째 가격이 정상가.
"""
from __future__ import annotations
import re, sys, io
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
PDF_DIR = ROOT / "pricepdf"
XLSX = ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx"

# Excel 차종 → PDF 파일 (한 차종이 여러 PDF 매칭 OK — 가솔린/하이브리드)
MODEL_TO_PDFS = {
    ("기아", "K5"):              ["price_k5.pdf"],
    ("기아", "K8"):              ["price_k8.pdf"],
    ("기아", "K9"):              ["price_k9.pdf"],
    ("기아", "모닝"):             ["price_morning.pdf"],
    ("기아", "니로"):             ["price_niro.pdf"],
    ("기아", "레이"):             ["price_ray.pdf"],
    ("기아", "셀토스"):            ["price_seltos.pdf"],
    ("기아", "쏘렌토"):            ["price_sorento.pdf"],
    ("기아", "스포티지"):           ["price_sportage.pdf"],
    ("기아", "카니발"):            ["price_carnival.pdf"],
    ("기아", "카니발 하이리무진"):     ["price_carnival-hi-limousine.pdf"],
    ("현대", "캐스퍼"):            ["AX_CASPER_price.pdf"],
    ("현대", "아반떼"):            ["avante-2026-price.pdf", "avante-hybrid-2026-price.pdf"],
    ("현대", "그랜저"):            ["the-new-grandeur-price.pdf", "the-new-grandeur-hybrid-price.pdf"],
    ("현대", "코나"):              ["kona-2027-price.pdf", "kona-hybrid-2027-price.pdf"],
    ("현대", "팰리세이드"):          ["palisade-2025-price.pdf", "palisade-hev-2025-price.pdf"],
    ("현대", "싼타페"):            ["santafe-2026-price.pdf", "santafe-hev-2026-price.pdf"],
    ("현대", "쏘나타"):            ["sonata-the-edge2026-price.pdf", "sonata-the-edge-hybrid2026-price.pdf"],
    ("현대", "투싼"):              ["tucson-2026-price.pdf", "tucson-hybrid-2026-price.pdf"],
    ("현대", "베뉴"):              ["venue-2025-price.pdf"],
    ("현대", "포터2"):             ["porter2-2026-price.pdf"],
    ("제네시스", "G70"):           ["genesis-g70-pricelist-kor-202605.pdf"],
    ("제네시스", "G80"):           ["genesis-g80-pricelist-kor-202605.pdf", "genesis-g80-black-pricelist-kor-202605.pdf"],
    ("제네시스", "G90"):           ["genesis-g90-pricelist-kor-202605.pdf", "genesis-g90-black-pricelist-kor-202605.pdf", "genesis-g90-long-wheel-base-black-pricelist-kor-202605.pdf"],
    ("제네시스", "GV70"):          ["genesis-gv70-pricelist-kor-202605.pdf"],
    ("제네시스", "GV80"):          ["genesis-gv80-pricelist-kor-202605.pdf"],
}


def load_excel():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['차량DB']
    rows = []
    for r in range(3, ws.max_row + 1):
        brand = ws.cell(r, 1).value
        model = ws.cell(r, 2).value
        detail = ws.cell(r, 4).value
        price = ws.cell(r, 5).value
        if not (brand and model and detail and price):
            continue
        rows.append({"brand": brand, "model": model, "detail": detail, "price": int(price), "row": r})
    return rows


def load_pdf_text(pdf_name: str) -> str:
    p = PDF_DIR / pdf_name.replace(".pdf", ".txt")
    return p.read_text(encoding='utf-8', errors='replace') if p.exists() else ""


def price_in_text(price: int, text: str) -> tuple[bool, str]:
    """가격이 PDF text 에 정확히 나오는지. 매칭한 패턴도 반환."""
    if not price or not text:
        return False, ""
    # 원 단위
    won_str = f"{price:,}"
    if won_str in text:
        return True, won_str
    if str(price) in text:
        return True, str(price)
    # 만 단위
    if price % 10_000 == 0:
        man = price // 10_000
        man_comma = f"{man:,}"
        # "3,393만" 형태
        pat = re.search(rf"{re.escape(man_comma)}\s*만", text)
        if pat:
            return True, pat.group()
        pat = re.search(rf"\b{re.escape(man_comma)}\b(?![만\d,])", text)
        if pat:
            return True, pat.group()
    return False, ""


def main():
    excel_rows = load_excel()
    print("# Excel 차량DB ↔ PDF 완전 일치 감사")
    print(f"- 검사 대상 Excel rows: **{len(excel_rows)}**")
    print()

    # 차종별 그룹화
    by_model = defaultdict(list)
    for row in excel_rows:
        by_model[(row["brand"], row["model"])].append(row)

    total = 0
    matched = 0
    missing_pdf = 0  # PDF 파일 자체가 매핑 안된 경우
    mismatch_rows = []

    for key, rows in sorted(by_model.items()):
        pdfs = MODEL_TO_PDFS.get(key)
        if not pdfs:
            missing_pdf += len(rows)
            print(f"### ⚠️ {key[0]} {key[1]} — PDF 매핑 없음 ({len(rows)} rows)")
            print()
            continue

        # 모든 PDF text 합치기 (가솔린+하이브리드 PDF 같이 검색)
        combined_text = "\n".join(load_pdf_text(p) for p in pdfs)
        model_mismatches = []
        for row in rows:
            total += 1
            ok, pat = price_in_text(row["price"], combined_text)
            if ok:
                matched += 1
            else:
                model_mismatches.append(row)
                mismatch_rows.append(row)

        if model_mismatches:
            print(f"### ❌ {key[0]} {key[1]} — {len(model_mismatches)}건 미매치 / 총 {len(rows)}")
            print(f"PDF: {', '.join(pdfs)}")
            print()
            print("| Excel row | 세부 트림 | Excel 가격 |")
            print("|---|---|---|")
            for r in model_mismatches:
                price_man = r["price"] / 10_000
                print(f"| r{r['row']} | {r['detail']} | {price_man:,.0f}만 ({r['price']:,}) |")
            print()

    print("---")
    print()
    print("## 종합 요약")
    print()
    print(f"- 검사한 Excel 트림: **{total}**")
    print(f"- ✅ PDF 매치: **{matched}** ({matched*100//total if total else 0}%)")
    print(f"- ❌ 미매치: **{total - matched}**")
    print(f"- ⚠️ PDF 매핑 누락 (모델 자체): **{missing_pdf}**")


if __name__ == "__main__":
    main()
