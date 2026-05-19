"""
Excel 차량DB 에 없는 ERP VEHICLE_DB 트림은 cascade 에서 숨김 (operating: false).
가격(만원) 기준으로 매칭. 매칭 안되면 'Excel 미운영' 으로 판단.

이전 작업 (apply_activate_excel_trims.py) 의 반대 방향.
"""
from __future__ import annotations
import re, sys, io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent

wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=True)
ws = wb["차량DB"]
excel_prices = set()
for r in range(3, ws.max_row + 1):
    p = ws.cell(r, 5).value
    if isinstance(p, (int, float)):
        excel_prices.add(int(p))

html_path = ROOT / "index.html"
html = html_path.read_text(encoding='utf-8')

# operating: false 가 없는 트림 entry 찾기, base_price_5 가 Excel set 에 없으면 false 추가
# 패턴: { trim_id: "...", name: "...", ..., base_price_5: NNNN, ...
# operating:false 있는 트림은 skip
pat = re.compile(
    r'(\{\s*trim_id:\s*"[^"]+"\s*,)(?!\s*operating:)(\s*name:\s*"([^"]+)"[^}]*?base_price_5:\s*(\d+))',
    re.DOTALL,
)

hidden = []
def replacer(m):
    head, body, name, p_man = m.group(1), m.group(2), m.group(3), int(m.group(4))
    if (p_man * 10_000) not in excel_prices:
        hidden.append((name, p_man))
        return head + ' operating: false,' + body
    return m.group(0)

new_html = pat.sub(replacer, html)

if hidden:
    html_path.write_text(new_html, encoding='utf-8')
    print(f"✅ {len(hidden)} 트림 숨김 (operating: false 추가)")
    for name, p in hidden[:40]:
        print(f"  {name} ({p:,}만)")
    if len(hidden) > 40:
        print(f"  (+ {len(hidden)-40} 건)")
else:
    print("변경 없음 (모든 트림이 Excel 에 있음 or 이미 숨김)")
