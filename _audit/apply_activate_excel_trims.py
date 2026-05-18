"""
Excel 차량DB 에 가격 일치하는 operating:false 트림 17건 → operating:false 제거.
"""
from __future__ import annotations
import re, sys, io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent

wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=True)
ws = wb["차량DB"]
excel_prices = set()
for r in range(3, ws.max_row + 1):
    p = ws.cell(r, 5).value
    if isinstance(p, (int, float)):
        excel_prices.add(int(p))

html_path = ROOT / "index.html"
html = html_path.read_text(encoding='utf-8')

# 트림 entry 전체 매치 (operating:false 포함)
pat = re.compile(
    r'(\{\s*trim_id:\s*"([^"]+)"\s*,)(\s*operating:\s*false\s*,)([^}]*?base_price_5:\s*(\d+))',
    re.DOTALL,
)

removed = 0
matches = []
def replacer(m):
    global removed
    head, tid, op_flag, body, p_man = m.group(1), m.group(2), m.group(3), m.group(4), int(m.group(5))
    if (p_man * 10_000) in excel_prices:
        removed += 1
        matches.append((tid, p_man))
        return head + body  # operating:false 빼기
    return m.group(0)

new_html = pat.sub(replacer, html)

if removed > 0:
    html_path.write_text(new_html, encoding='utf-8')
    print(f"✅ {removed} 트림 활성화 (operating:false 제거)")
    for tid, p in matches:
        print(f"  {tid} ({p:,}만)")
else:
    print("변경 없음 (Excel 매치 트림 없음 or 이미 활성)")
