"""Excel 의 모든 시트에서 차량DB col P (인수가적용) 참조 찾기."""
import openpyxl, re, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
from pathlib import Path
ROOT = Path(__file__).resolve().parent.parent
wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=False)

# col P 또는 13 returning VLOOKUP 검색
P_PAT = re.compile(r"차량DB!\$?P|,\s*13\s*[,)]")
for sname in wb.sheetnames:
    ws = wb[sname]
    hits = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "VLOOKUP" in cell.value and "차량DB" in cell.value:
                if P_PAT.search(cell.value):
                    hits.append((cell.coordinate, cell.value[:250]))
    if hits:
        print(f"=== {sname} (col P 사용) ===")
        for a, v in hits[:8]:
            print(f"  {a}: {v}")
        print()
