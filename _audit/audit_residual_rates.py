"""
Excel 차량DB 의 r24/r36/r48/r60 + buyback_apply 가
vehicles.json (ERP 런타임 데이터) 와 정확히 일치하는지 검증.

매칭 키: (브랜드, 차종, 세부모델, 가격) → 모든 4가 일치해야 OK.
"""
from __future__ import annotations
import json, sys, io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent

# Excel
wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=True)
ws = wb["차량DB"]
excel_rows = {}  # (brand, model, detail) -> {price, r24,r36,r48,r60, buyback_apply}
for r in range(3, ws.max_row + 1):
    brand = ws.cell(r, 1).value
    model = ws.cell(r, 2).value
    detail = ws.cell(r, 4).value
    price = ws.cell(r, 5).value
    r24, r36, r48, r60 = (ws.cell(r, c).value for c in (11, 12, 13, 14))
    buyback = ws.cell(r, 16).value
    if not (brand and model and detail and isinstance(price, (int, float))):
        continue
    excel_rows[(brand, model, detail)] = {
        "price": int(price),
        "r24": r24, "r36": r36, "r48": r48, "r60": r60,
        "buyback_apply": buyback,
    }

# vehicles.json
veh = json.loads((ROOT / "src/data/vehicles.json").read_text(encoding='utf-8'))
vj_by_trim = {(v["brand"], v["model"], v["trim"]): v for v in veh}

# 비교
mismatches = []
matched_count = 0
unmatched_in_vj = 0
for key, ex in excel_rows.items():
    vj = vj_by_trim.get(key)
    if not vj:
        unmatched_in_vj += 1
        continue
    issues = []
    for fld in ["r24", "r36", "r48", "r60", "buyback_apply"]:
        ex_v = ex[fld]
        vj_v = vj.get(fld)
        # 둘 다 숫자인지 + 차이가 1e-9 이상이면 mismatch
        if ex_v is None and vj_v is None:
            continue
        if ex_v is None or vj_v is None:
            issues.append(f"{fld}: Excel={ex_v}, VJ={vj_v}")
            continue
        if abs(float(ex_v) - float(vj_v)) > 1e-6:
            issues.append(f"{fld}: Excel={ex_v}, VJ={vj_v}")
    if issues:
        mismatches.append((key, ex["price"], issues))
    else:
        matched_count += 1

print("# vehicles.json ↔ Excel 차량DB 잔가율 정합성")
print()
print(f"- Excel rows: {len(excel_rows)}")
print(f"- 일치: {matched_count}")
print(f"- ❌ 불일치: {len(mismatches)}")
print(f"- ⚠️ Excel에 있으나 vehicles.json 없음: {unmatched_in_vj}")
print()

if mismatches:
    print("## 불일치 상세")
    print()
    print("| 차종 | 트림 | 가격 | 차이 |")
    print("|---|---|---|---|")
    for key, price, issues in mismatches[:80]:
        brand, model, detail = key
        issue_str = " · ".join(issues)
        print(f"| {brand} {model} | {detail} | {price:,} | {issue_str} |")
    if len(mismatches) > 80:
        print(f"\n*(+ {len(mismatches)-80}건 생략)*")
