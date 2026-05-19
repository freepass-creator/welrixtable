"""
같은 차종 + 같은 엔진(연료/배기량) 인데 트림에 따라 잔가율이 다른 경우 찾기.
원칙: 잔가율은 트림과 무관해야 함 (같은 자동차 본체).
"""
from __future__ import annotations
import sys, io
from pathlib import Path
from collections import defaultdict
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent
wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=True)
ws = wb["차량DB"]

# 그룹화 키: (브랜드, 차종, 유종, 배기량) — 같은 본체 차량 그룹
groups = defaultdict(list)
for r in range(3, ws.max_row + 1):
    brand = ws.cell(r, 1).value
    model = ws.cell(r, 2).value
    detail = ws.cell(r, 4).value
    disp = ws.cell(r, 6).value
    fuel = ws.cell(r, 7).value
    rates = (ws.cell(r, 11).value, ws.cell(r, 12).value, ws.cell(r, 13).value, ws.cell(r, 14).value)
    if not (brand and model and detail and fuel and disp):
        continue
    groups[(brand, model, fuel, disp)].append({
        "detail": detail, "rates": rates, "row": r,
    })

print("# 같은 차종·엔진 내 트림간 잔가율 불일치")
print()
anomalies = 0
for key, items in groups.items():
    rate_sets = set(it["rates"] for it in items)
    if len(rate_sets) > 1:
        anomalies += 1
        brand, model, fuel, disp = key
        print(f"### {brand} {model} {fuel} {disp}cc — 트림별 잔가율 분기")
        print()
        print("| Excel row | 트림 | r24 | r36 | r48 | r60 |")
        print("|---|---|---|---|---|---|")
        for it in items:
            d = it["detail"]
            r24, r36, r48, r60 = it["rates"]
            mark = ""
            # 다른 트림과 가장 다른 값을 강조
            print(f"| r{it['row']} | {d} | {r24} | {r36} | {r48} | {r60} |")
        print()

print(f"\n불일치 그룹: **{anomalies}**개")
