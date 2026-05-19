"""
견적1 시트의 모든 formula 셀을 grid 형태로 dump.
- 빈 셀 제외
- 각 셀의 formula + 캐시된 value 같이
- A~AZ 컬럼, 1~131 행 (Excel 견적1 max_row = 131)
"""
import openpyxl, sys, io
from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent

XLSX = ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx"

wb_f = openpyxl.load_workbook(XLSX, data_only=False)
wb_v = openpyxl.load_workbook(XLSX, data_only=True)

def get_letter(c):
    return openpyxl.utils.get_column_letter(c)

ws_f = wb_f['견적1']
ws_v = wb_v['견적1']

print("# 견적1 시트 — 전체 cell formula + value")
print()
print(f"max_row={ws_f.max_row}, max_col={ws_f.max_column}")
print()

# Section 1: 모든 formula cell (=로 시작) 행 단위로
print("## Formula cells (수식 있는 셀)")
print()
print("| Addr | Value | Formula |")
print("|---|---|---|")
for r in range(1, ws_f.max_row + 1):
    for c in range(1, ws_f.max_column + 1):
        v = ws_f.cell(r, c).value
        if isinstance(v, str) and v.startswith('='):
            val = ws_v.cell(r, c).value
            val_str = str(val) if val is not None else '—'
            if isinstance(val, float):
                val_str = f"{val:.4f}".rstrip('0').rstrip('.')
            addr = f"{get_letter(c)}{r}"
            # formula 너무 길면 잘라서
            f_short = v[:300]
            print(f"| {addr} | {val_str} | `{f_short}` |")
print()

# Section 2: 상수 (formula 아니지만 값 있는 셀)
print("## 상수/입력 셀 (formula 아님, 값 있음)")
print()
print("| Addr | Value |")
print("|---|---|")
for r in range(1, ws_f.max_row + 1):
    for c in range(1, ws_f.max_column + 1):
        v = ws_f.cell(r, c).value
        if v is None:
            continue
        if isinstance(v, str) and v.startswith('='):
            continue
        addr = f"{get_letter(c)}{r}"
        v_str = str(v)
        if isinstance(v, float):
            v_str = f"{v:.6f}".rstrip('0').rstrip('.')
        if len(v_str) > 80:
            v_str = v_str[:80] + '...'
        print(f"| {addr} | {v_str} |")
