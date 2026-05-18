"""
ERP VEHICLE_DB (index.html 인라인) 에서 operating: false 처리된 트림 중
Excel 차량DB 에 가격이 일치(±5만원)하는 트림 찾기.
→ Excel에 있는데 ERP UI에서 숨겨진 트림 = 활성화해야 할 후보.
"""
from __future__ import annotations
import re, sys, io
from pathlib import Path
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
ROOT = Path(__file__).resolve().parent.parent

# Excel 가격 set (원 단위)
wb = openpyxl.load_workbook(ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx", data_only=True)
ws = wb["차량DB"]
excel_prices = {}  # price_won → 트림 디테일
for r in range(3, ws.max_row + 1):
    p = ws.cell(r, 5).value
    detail = ws.cell(r, 4).value
    if isinstance(p, (int, float)) and detail:
        excel_prices[int(p)] = detail

# ERP VEHICLE_DB 에서 operating: false + base_price_5 매치
html = (ROOT / "index.html").read_text(encoding='utf-8')

# trim line 패턴: { trim_id: "xxx", operating: false, name: "yyy", ..., base_price_5: 1460, ...
pat = re.compile(r'\{\s*trim_id:\s*"([^"]+)"[^}]*?operating:\s*false[^}]*?name:\s*"([^"]+)"[^}]*?base_price_5:\s*(\d+)', re.DOTALL)

print("# operating:false 트림 중 Excel 가격 일치 (활성화 후보)")
print()
print(f"- Excel 가격 set: {len(excel_prices)}건")
print(f"- ERP operating:false 검사 중...")
print()
print("| trim_id | name | price(만원) | Excel detail |")
print("|---|---|---|---|")

found = 0
for m in pat.finditer(html):
    tid, name, p_man = m.group(1), m.group(2), int(m.group(3))
    p_won = p_man * 10_000
    if p_won in excel_prices:
        found += 1
        print(f"| {tid} | {name} | {p_man:,} | {excel_prices[p_won]} |")

print(f"\n매치: **{found}** / 전체 operating:false")
