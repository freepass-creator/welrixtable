"""
v3: 3-source 정합성 감사 — Catalog (.json) vs Excel (차량DB sheet) vs PDF (.txt)
PDF가 정답. 카탈로그/엑셀이 PDF에 매칭되는지 확인.

전략: Excel은 차명상세 텍스트가 풍부하므로 PDF base price와 직접 매칭하기 쉽다.
Catalog는 PDF base price와 매칭. Excel과 Catalog 가격 set 비교.

출력: 모델별 3-source price set diff table.
"""
from __future__ import annotations
import json, re, sys, io
from pathlib import Path
from collections import defaultdict

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

ROOT = Path(__file__).resolve().parent.parent
PDF_DIR = ROOT / "pricepdf"
CATALOG_DIR = ROOT / "src" / "data" / "car-master"
XLSX = ROOT / "웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx"

# ============================================================
# Excel: 차량DB 시트 → {(브랜드, 차종) -> [(trim_detail, 가격)]}
# ============================================================
def load_excel_db() -> dict[tuple[str, str], list[tuple[str, int]]]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['차량DB']
    out: dict = defaultdict(list)
    for r in range(3, ws.max_row + 1):
        brand = ws.cell(r, 1).value
        model = ws.cell(r, 2).value
        trim_detail = ws.cell(r, 4).value
        price = ws.cell(r, 5).value
        if not (brand and model and trim_detail and price):
            continue
        out[(brand, model)].append((trim_detail, int(price)))
    return out


# ============================================================
# Catalog → {catalog_id -> {trim_name: base_won}}
# ============================================================
def load_catalogs() -> dict[str, dict[str, int]]:
    out = {}
    for p in CATALOG_DIR.glob("*.json"):
        if p.stem.startswith("_"):
            continue
        try:
            cat = json.loads(p.read_text(encoding='utf-8'))
        except Exception:
            continue
        trims = cat.get("trims", {})
        out[p.stem] = {n: info.get("price", {}).get("base", 0) for n, info in trims.items()}
    return out


# ============================================================
# PDF (text) → text body
# ============================================================
def load_pdf_text(pdf_name: str) -> str:
    p = PDF_DIR / pdf_name.replace(".pdf", ".txt")
    return p.read_text(encoding='utf-8', errors='replace') if p.exists() else ""


def price_in_text(price: int, text: str) -> bool:
    if not price or not text:
        return False
    if f"{price:,}" in text or str(price) in text:
        return True
    man = price // 10_000
    if re.search(rf"{man:,}\s*만", text) or re.search(rf"{man}\s*만", text):
        return True
    return False


# ============================================================
# 매핑: (Excel 차종) ↔ (PDF 파일) ↔ (catalog id list)
# ============================================================
MODEL_GROUPS = [
    # (excel_brand, excel_model_kr, pdf_name, [catalog_ids])
    ("기아", "K5",        "price_k5.pdf",          ["kia_k5_dl3_facelift", "kia_k5_dl3_facelift_hybrid"]),
    ("기아", "K8",        "price_k8.pdf",          ["kia_k8_gl3", "kia_k8_gl3_hybrid", "kia_k8_gl3_pre", "kia_k8_gl3_pre_hybrid"]),
    ("기아", "K9",        "price_k9.pdf",          ["kia_k9_rj", "kia_k9_rj_pre"]),
    ("기아", "모닝",       "price_morning.pdf",     ["kia_morning_sa_facelift", "kia_morning_sa", "kia_morning_ja"]),
    ("기아", "니로",       "price_niro.pdf",        ["kia_niro_sg2", "kia_niro_sg2_hybrid"]),
    ("기아", "레이",       "price_ray.pdf",         ["kia_ray_tam_facelift1", "kia_ray", "kia_ray_ev"]),
    ("기아", "셀토스",      "price_seltos.pdf",      ["kia_seltos_sp3", "kia_seltos_sp2", "kia_seltos"]),
    ("기아", "쏘렌토",      "price_sorento.pdf",     ["kia_sorento_mq4_facelift", "kia_sorento_mq4_facelift_hybrid", "kia_sorento_mq4", "kia_sorento_mq4_hybrid"]),
    ("기아", "스포티지",     "price_sportage.pdf",    ["kia_sportage_nq5", "kia_sportage_nq5_hybrid"]),
    ("기아", "카니발",      "price_carnival.pdf",    ["kia_carnival_ka4_facelift", "kia_carnival_ka4_facelift_hybrid", "kia_carnival_ka4"]),
    ("기아", "카니발 하이리무진","price_carnival-hi-limousine.pdf",["kia_carnival_ka4_hi_limousine_facelift", "kia_carnival_ka4_hi_limousine"]),

    ("현대", "캐스퍼",      "AX_CASPER_price.pdf",   ["hyundai_casper", "hyundai_casper_pre"]),
    ("현대", "아반떼",      "avante-2026-price.pdf", ["hyundai_avante_cn7", "hyundai_avante_cn7_pre"]),
    ("현대", "아반떼",      "avante-hybrid-2026-price.pdf", ["hyundai_avante_cn7_hybrid"]),
    ("현대", "그랜저",      "the-new-grandeur-price.pdf", ["hyundai_grandeur_gn7_fl", "hyundai_grandeur_gn7"]),
    ("현대", "그랜저",      "the-new-grandeur-hybrid-price.pdf", ["hyundai_grandeur_gn7_fl_hybrid", "hyundai_grandeur_gn7_hybrid"]),
    ("현대", "코나",        "kona-2027-price.pdf",   ["hyundai_kona_sx2"]),
    ("현대", "코나",        "kona-hybrid-2027-price.pdf", ["hyundai_kona_sx2_hybrid"]),
    ("현대", "팰리세이드",    "palisade-2025-price.pdf", ["hyundai_palisade_lx3"]),
    ("현대", "팰리세이드",    "palisade-hev-2025-price.pdf", ["hyundai_palisade_lx3_hybrid"]),
    ("현대", "싼타페",      "santafe-2026-price.pdf", ["hyundai_santafe_mx5"]),
    ("현대", "싼타페",      "santafe-hev-2026-price.pdf", ["hyundai_santafe_mx5_hybrid"]),
    ("현대", "쏘나타",      "sonata-the-edge2026-price.pdf", ["hyundai_sonata_dn8_edge"]),
    ("현대", "쏘나타",      "sonata-the-edge-hybrid2026-price.pdf", ["hyundai_sonata_dn8_edge_hybrid"]),
    ("현대", "투싼",        "tucson-2026-price.pdf", ["hyundai_tucson_nx4"]),
    ("현대", "투싼",        "tucson-hybrid-2026-price.pdf", ["hyundai_tucson_nx4_hybrid"]),
    ("현대", "베뉴",        "venue-2025-price.pdf",  ["hyundai_venue"]),

    ("제네시스", "G70",     "genesis-g70-pricelist-kor-202605.pdf", ["genesis_g70", "genesis_g70_st"]),
    ("제네시스", "G80",     "genesis-g80-pricelist-kor-202605.pdf", ["genesis_g80_rg3", "genesis_g80_rg3_pre"]),
    ("제네시스", "G90",     "genesis-g90-pricelist-kor-202605.pdf", ["genesis_g90_rs4"]),
    ("제네시스", "GV70",    "genesis-gv70-pricelist-kor-202605.pdf", ["genesis_gv70", "genesis_gv70_e", "genesis_gv70_pre"]),
    ("제네시스", "GV80",    "genesis-gv80-pricelist-kor-202605.pdf", ["genesis_gv80", "genesis_gv80_coupe", "genesis_gv80_pre"]),
]


def fmt_man(w: int | None) -> str:
    if not w:
        return "—"
    return f"{w/10_000:,.0f}만"


def main():
    excel_db = load_excel_db()
    catalogs = load_catalogs()

    print("# 3-Source 정합성 감사 (Excel · PDF · Catalog)")
    print()
    print("- **기준**: PDF (제조사 공식 가격표)")
    print("- **개소세**: 5% 정상가 기준")
    print()

    grand_total = {"excel_only": 0, "catalog_only": 0, "both": 0, "all_three": 0}

    for brand, model, pdf_name, cat_ids in MODEL_GROUPS:
        excel_trims = excel_db.get((brand, model), [])
        pdf_text = load_pdf_text(pdf_name)

        # 카탈로그 합산
        cat_trims = []  # [(catalog_id, trim_name, price)]
        for cid in cat_ids:
            for n, p in catalogs.get(cid, {}).items():
                cat_trims.append((cid, n, p))

        # 가격 set 만들기 (만원 단위로 정규화)
        excel_prices = {p // 10_000: t for t, p in excel_trims}
        cat_prices = {p // 10_000: f"{cid}:{n}" for cid, n, p in cat_trims if p}

        # 어느쪽에만 있는지
        excel_set = set(excel_prices.keys())
        cat_set = set(cat_prices.keys())

        excel_only_set = excel_set - cat_set
        cat_only_set = cat_set - excel_set
        both = excel_set & cat_set

        # PDF 매칭 (만원 set 중 PDF text에 있는 것)
        pdf_match = {p for p in excel_set | cat_set if p and price_in_text(p * 10_000, pdf_text)}

        grand_total["excel_only"] += len(excel_only_set)
        grand_total["catalog_only"] += len(cat_only_set)
        grand_total["both"] += len(both)
        grand_total["all_three"] += len(both & pdf_match)

        # 출력 — 문제 있는 경우만
        has_issues = excel_only_set or cat_only_set or (both - pdf_match)
        if not has_issues:
            continue

        print(f"### {brand} {model}  —  `{pdf_name}`")
        print()
        print(f"- Excel trims: {len(excel_trims)}건 / Catalog trims: {len(cat_trims)}건")
        print()

        if cat_only_set:
            print(f"**🔴 Catalog 에만 있는 가격** (Excel·PDF 어느쪽에도 없음 → 잘못된 트림 가능성)")
            print()
            print("| 카탈로그 | 트림 | 가격 | PDF 매치 |")
            print("|---|---|---|---|")
            for p in sorted(cat_only_set):
                cinfo = cat_prices[p]
                pdf_ok = "✓" if p in pdf_match else "❌"
                cid, tname = cinfo.split(":", 1)
                print(f"| `{cid}` | {tname} | {fmt_man(p*10_000)} | {pdf_ok} |")
            print()

        if excel_only_set:
            print(f"**🟡 Excel 에만 있는 가격** (Catalog 누락 → 추가 필요할 수 있음)")
            print()
            print("| Excel 차명상세 | 가격 | PDF 매치 |")
            print("|---|---|---|")
            for p in sorted(excel_only_set):
                tname = excel_prices[p]
                pdf_ok = "✓" if p in pdf_match else "❌"
                print(f"| {tname} | {fmt_man(p*10_000)} | {pdf_ok} |")
            print()

        both_no_pdf = both - pdf_match
        if both_no_pdf:
            print(f"**⚠️  Excel·Catalog 일치하지만 PDF에 없음** (3 소스 모두 stale일 수도)")
            print()
            print("| 가격 | Excel 트림 |")
            print("|---|---|")
            for p in sorted(both_no_pdf):
                print(f"| {fmt_man(p*10_000)} | {excel_prices[p]} |")
            print()

        print("---")
        print()

    print("## 종합 요약")
    print()
    print(f"- 3소스 모두 일치: **{grand_total['all_three']}**개 트림 (이상적)")
    print(f"- Excel∩Catalog 일치: **{grand_total['both']}**개 (그 중 PDF도 일치: {grand_total['all_three']}개)")
    print(f"- 🟡 Excel only (Catalog 누락): **{grand_total['excel_only']}**개")
    print(f"- 🔴 Catalog only (Excel 누락): **{grand_total['catalog_only']}**개")


if __name__ == "__main__":
    main()
