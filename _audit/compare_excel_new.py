import openpyxl, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

NEW = r'C:\dev\welrixtable\웰릭스모빌리티 신차 견적(20260506)_v4.5_배포용.xlsx'
wb = openpyxl.load_workbook(NEW, data_only=True)

print('SHEETS:')
for i, s in enumerate(wb.sheetnames):
    print(f'  [{i}] {s}')
print()

# 차량DB sheet — find by name containing "차량" or "DB"
db_idx = None
for i, s in enumerate(wb.sheetnames):
    if '차량' in s or 'DB' in s.upper():
        db_idx = i; break

ws = wb.worksheets[db_idx if db_idx is not None else 9]
print(f'=== Vehicle DB sheet: {ws.title} ({ws.max_row} rows) ===')

# Print header
hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
print('HEADER:', hdr)
print()

# Find Palisade
print('=== 팰리세이드 rows ===')
for r in range(2, ws.max_row + 1):
    row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    joined = ' | '.join(str(v) for v in row)
    if '팰리세이드' in joined:
        # Compact: brand, trim, price, r24..r60, strategic, buyback
        print(f'R{r}:', row)
